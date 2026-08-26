#!/usr/bin/env python3
"""Extract UN WPP 2024 GEN/01 indicators into small JSON shards for git.

Reads data/WPP2024_GEN_F01_DEMOGRAPHIC_INDICATORS_FULL.xlsx
Writes public/data/wpp2024/{manifest,tfr,population,e0,net-migration,births,median-age,srb}.json
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "WPP2024_GEN_F01_DEMOGRAPHIC_INDICATORS_FULL.xlsx"
OUT = ROOT / "public" / "data" / "wpp2024"

WANTED = {
    "iso3": ("ISO3 Alpha-code",),
    "type": ("Type",),
    "year": ("Year",),
    "pop": ("Total Population, as of 1 July (thousands)",),
    "tfr": ("Total Fertility Rate (live births per woman)",),
    "e0": ("Life Expectancy at Birth, both sexes (years)",),
    "mig": ("Net Number of Migrants (thousands)",),
    "births": ("Births (thousands)",),
    "median": ("Median Age, as of 1 July (years)",),
    "srb": ("Sex Ratio at Birth (males per 100 female births)",),
}

SHARDS = [
    ("tfr", "tfr", 1.0, 4),
    ("population", "pop", 1000.0, 0),
    ("e0", "e0", 1.0, 3),
    ("net-migration", "mig", 1000.0, 0),
    ("births", "births", 1000.0, 0),
    ("median-age", "median", 1.0, 2),
    ("srb", "srb", 0.01, 4),  # males/100 females → boys per girl
]


def header_map(row) -> dict[str, int]:
    found = {}
    for i, cell in enumerate(row):
        if cell is None:
            continue
        text = str(cell).strip()
        for key, names in WANTED.items():
            if text in names:
                found[key] = i
    return found


def to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        if v != v:  # NaN
            return None
        return float(v)
    s = str(v).strip().replace(",", "")
    if s in ("...", "—", "-", ""):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def pack(value: float, scale: float, digits: int):
    x = value * scale
    if digits <= 0:
        return int(round(x))
    return round(x, digits)


def ingest_sheet(ws, cols, bucket: str, out: dict):
    """bucket is 'e' (estimates) or 'm' (medium)."""
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        typ = row[cols["type"]] if cols["type"] < len(row) else None
        if typ != "Country/Area":
            continue
        iso3 = row[cols["iso3"]]
        year = row[cols["year"]]
        if not iso3 or year is None:
            continue
        iso3 = str(iso3).strip()
        try:
            year = int(year)
        except (TypeError, ValueError):
            continue
        rec = out.setdefault(iso3, {})
        for shard, key, scale, digits in SHARDS:
            v = to_float(row[cols[key]] if cols[key] < len(row) else None)
            if v is None:
                continue
            rec.setdefault(shard, {}).setdefault(bucket, []).append([year, pack(v, scale, digits)])
        n += 1
    return n


def main():
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Installing openpyxl…", file=sys.stderr)
        import subprocess

        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        from openpyxl import load_workbook

    if not XLSX.exists():
        raise SystemExit(f"Missing {XLSX}")

    print("Opening", XLSX.name, f"({XLSX.stat().st_size / 1e6:.1f} MB)")
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    print("Sheets:", wb.sheetnames[:6], "…")

    # Discover header on Estimates
    est = wb["Estimates"]
    cols = None
    header_row = 1
    peek = est.iter_rows(min_row=1, max_row=20, values_only=True)
    for i, row in enumerate(peek, start=1):
        m = header_map(row)
        if "iso3" in m and "tfr" in m and "year" in m and "type" in m:
            cols = m
            header_row = i
            break
    if not cols:
        raise SystemExit("Could not find indicator header row")
    print("Header row", header_row, "cols", cols)

    # reopen because we consumed the iterator; read_only worksheets are one-pass
    wb.close()
    wb = load_workbook(XLSX, read_only=True, data_only=True)

    data: dict[str, dict] = {}
    est = wb["Estimates"]
    # skip until after header
    rows = est.iter_rows(min_row=header_row + 1, values_only=True)

    def ingest_iter(rows, bucket):
        n = 0
        for row in rows:
            typ = row[cols["type"]] if cols["type"] < len(row) else None
            if typ != "Country/Area":
                continue
            iso3 = row[cols["iso3"]]
            year = row[cols["year"]]
            if not iso3 or year is None:
                continue
            iso3 = str(iso3).strip()
            try:
                year = int(year)
            except (TypeError, ValueError):
                continue
            rec = data.setdefault(iso3, {})
            for shard, key, scale, digits in SHARDS:
                v = to_float(row[cols[key]] if cols[key] < len(row) else None)
                if v is None:
                    continue
                rec.setdefault(shard, {}).setdefault(bucket, []).append([year, pack(v, scale, digits)])
            n += 1
        return n

    n_est = ingest_iter(rows, "e")
    print(f"Estimates country-years: {n_est}")

    med = wb["Medium variant"]
    med_header = 1
    for i, row in enumerate(med.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        m = header_map(row)
        if "iso3" in m and "tfr" in m:
            med_header = i
            break
    wb.close()
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    med = wb["Medium variant"]
    n_med = ingest_iter(med.iter_rows(min_row=med_header + 1, values_only=True), "m")
    print(f"Medium country-years: {n_med}")
    wb.close()

    OUT.mkdir(parents=True, exist_ok=True)
    manifest = {
        "source": "United Nations, Department of Economic and Social Affairs, Population Division (2024). World Population Prospects 2024.",
        "license": "CC BY 3.0 IGO",
        "file": "GEN/01 Demographic indicators",
        "variants": ["Estimates 1950-2023", "Medium 2024-2100"],
        "countries": len(data),
        "shards": [],
    }

    for shard, _key, _s, _d in SHARDS:
        payload = {}
        for iso3, rec in data.items():
            series = rec.get(shard)
            if not series:
                continue
            for bucket in ("e", "m"):
                pts = series.get(bucket)
                if pts:
                    pts.sort(key=lambda p: p[0])
            payload[iso3] = series
        path = OUT / f"{shard}.json"
        path.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")
        size = path.stat().st_size
        manifest["shards"].append({"id": shard, "file": f"{shard}.json", "bytes": size})
        print(f"  {path.name:20s} {size/1024:8.1f} KB")
        if size > 2_000_000:
            print("  WARNING: shard exceeds 2 MB; split by ISO3 prefix if needed")

    (OUT / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
